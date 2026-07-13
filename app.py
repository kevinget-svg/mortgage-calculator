import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="房贷计算器",
    page_icon="🏠",
    layout="wide",
    menu_items={"About": "房贷计算器 — 支持首套/二套、公积金/商贷/组合贷"},
)
st.markdown("""<style>
    /* 移动端触控优化 */
    @media (max-width: 768px) {
        .stButton button { min-height: 44px; font-size: 16px; }
        .stSelectbox div[data-baseweb="select"] { font-size: 16px; }
    }
</style>""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────
PRICE_OPTIONS = [50, 80, 100, 120, 150, 180, 200, 220, 250, 280, 300, 350, 400, 450, 500]
PRICE_DEFAULTS = [150, 200, 250]
YEAR_OPTIONS = [5, 10, 15, 20, 25, 30]

# ── Helpers ────────────────────────────────────────────────────────────────
def monthly_payment(principal: float, annual_rate: float, years: int):
    """等额本息：返回 (月供, 总利息, 还款总额)，单位均为元"""
    if principal <= 0 or years <= 0:
        return 0.0, 0.0, 0.0
    r = annual_rate / 100 / 12
    n = years * 12
    if r == 0:
        m = principal / n
    else:
        m = principal * r * (1 + r) ** n / ((1 + r) ** n - 1)
    total = m * n
    interest = total - principal
    return m, interest, total


def make_scenario_label(price: float, dp: int):
    """生成场景标签，如 '200万/20%'"""
    return f"{price:.0f}万/{dp}%"


def _parse_rgb(color: str):
    """将 '#hex' 或 'rgb(r,g,b)' 解析为 (r,g,b)"""
    color = color.strip()
    if color.startswith('#'):
        color = color.lstrip('#')
        return int(color[:2], 16), int(color[2:4], 16), int(color[4:], 16)
    if color.startswith('rgb'):
        vals = color.replace('rgb(', '').replace(')', '').split(',')
        return int(vals[0]), int(vals[1]), int(vals[2])
    raise ValueError(f"Unsupported color format: {color}")


def lighten(color: str, factor: float = 0.4) -> str:
    """与白色混合，返回更浅的 hex 颜色"""
    r, g, b = _parse_rgb(color)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'#{r:02x}{g:02x}{b:02x}'


def darken(color: str, factor: float = 0.4) -> str:
    """与黑色混合，返回更深的 hex 颜色"""
    r, g, b = _parse_rgb(color)
    r = int(r * (1 - factor))
    g = int(g * (1 - factor))
    b = int(b * (1 - factor))
    return f'#{r:02x}{g:02x}{b:02x}'


def build_excel(df: pd.DataFrame, loan_type: str, years: int) -> bytes:
    """用 openpyxl 生成带格式的 Excel 文件，返回 bytes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "贷款对比结果"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1, value=f"房贷计算结果 — {loan_type}，{years}年")
    title_cell.font = Font(name="微软雅黑", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")
    ws.append([])
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for row_idx, row in df.iterrows():
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx + 4, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            len(str(df.columns[col_idx - 1])),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, 12)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ── Sidebar ────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 基本设置")
    property_type = st.radio("房产类型", ["首套房", "二套房"], horizontal=True)
    loan_type = st.radio("贷款方式", ["纯商业贷款", "纯公积金贷款", "组合贷款"])
    years = st.selectbox("贷款年限", YEAR_OPTIONS, index=5)

# ── 利率 & 规则 ────────────────────────────────────────────────────────────
gjj_rate = 2.6
sd_rate = 3.0 if property_type == "首套房" else 3.3
uses_gjj = loan_type in ("纯公积金贷款", "组合贷款")
min_down_pay = 20 if uses_gjj else 15

st.markdown(
    f"**公积金利率** `{gjj_rate}%` ｜ "
    f"**商贷利率** `{sd_rate}%` ｜ "
    f"**最低首付** `{min_down_pay}%`"
)

# ── 自定义输入解析 ─────────────────────────────────────────────────────────
def parse_custom_input(text: str):
    """解析逗号/空格/换行分隔的数字字符串，返回 float 列表"""
    if not text or not text.strip():
        return []
    text = text.replace("\n", " ").replace(",", " ").replace("，", " ")
    result = []
    for part in text.split():
        try:
            result.append(float(part))
        except ValueError:
            pass
    return result


# ═══════════════════════════════════════════════════════════════════════════════
tab1, tab2 = st.tabs(["📊 方案对比", "🔄 首套二套顺序对比"])

# ═══════════════════════════════════════════════════════════════════════════════
#  Tab 1: 多选方案对比
# ═══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("📝 选择对比场景")
    col1, col2, col3 = st.columns(3)

    with col1:
        selected_prices = st.multiselect(
            "房屋总价（万元）",
            options=PRICE_OPTIONS,
            default=[p for p in PRICE_DEFAULTS if p in PRICE_OPTIONS],
            help="可多选，同时对比不同总价",
        )
        with st.expander("📝 自定义总价"):
            custom_prices_str = st.text_input(
                "输入数值（逗号/空格/换行分隔）",
                placeholder="例如：175 225 275",
                help="会与上方多选结果合并",
                label_visibility="collapsed",
            )
            custom_prices = parse_custom_input(custom_prices_str)

    with col2:
        dp_pcts = list(range(min_down_pay, 55, 5))
        default_dp = [min_down_pay, min_down_pay + 5, min_down_pay + 10]
        default_dp = [d for d in default_dp if d in dp_pcts]
        selected_dps = st.multiselect(
            "首付比例（%）",
            options=dp_pcts,
            default=default_dp[:3],
            help="可多选，同时对比不同首付比例",
        )
        with st.expander("📝 自定义首付"):
            custom_dps_str = st.text_input(
                "输入数值（逗号/空格/换行分隔）",
                placeholder="例如：18 22 28",
                help="会与上方多选结果合并",
                label_visibility="collapsed",
            )
            custom_dps = parse_custom_input(custom_dps_str)

    with col3:
        max_gjj_wan = 80.0
        if uses_gjj:
            max_gjj_wan = st.number_input(
                "公积金最高贷款额度（万元）",
                min_value=0.0, value=80.0, step=1.0, format="%.0f",
                help="各地公积金中心规定的最高可贷额度",
            )

    # 合并自定义值
    all_prices = sorted(set(list(selected_prices) + custom_prices))
    all_dps = sorted(set(list(selected_dps) + custom_dps))

    if not all_prices or not all_dps:
        st.info("👆 请至少选择一个房屋总价和首付比例（多选或自定义皆可）。")
    else:
        # ── 计算所有组合 ───────────────────────────────────────────────────────
        rows = []
        selected_prices = all_prices
        selected_dps = all_dps
        max_gjj = max_gjj_wan * 10000

        for price_wan in selected_prices:
            total_price = price_wan * 10000
            for dp_pct in selected_dps:
                down_payment = total_price * dp_pct / 100
                total_loan = total_price - down_payment

                if loan_type == "纯公积金贷款":
                    gjj_loan = min(total_loan, max_gjj) if max_gjj > 0 else total_loan
                    sd_loan = 0.0
                    gjj_note = "⚠已达上限" if (total_loan > max_gjj and max_gjj > 0) else ""
                elif loan_type == "纯商业贷款":
                    gjj_loan = 0.0
                    sd_loan = total_loan
                    gjj_note = ""
                else:
                    gjj_loan = min(total_loan, max_gjj) if max_gjj > 0 else 0
                    sd_loan = total_loan - gjj_loan
                    gjj_note = "⚠已达上限" if (max_gjj > 0 and gjj_loan >= max_gjj) else ""

                gjj_m, gjj_interest, gjj_total = monthly_payment(gjj_loan, gjj_rate, years)
                sd_m, sd_interest, sd_total = monthly_payment(sd_loan, sd_rate, years)

                row = {
                    "场景": make_scenario_label(price_wan, dp_pct),
                    "房屋总价(万)": price_wan,
                    "首付比例": f"{dp_pct}%",
                    "首付(万)": round(down_payment / 10000, 2),
                    "公积金贷款(万)": f"{gjj_loan/10000:.2f}{gjj_note}" if gjj_loan > 0 else 0,
                    "公积金月供(元)": round(gjj_m, 0) if gjj_loan > 0 else 0,
                    "公积金总利息(万)": round(gjj_interest / 10000, 2) if gjj_loan > 0 else 0,
                    "商贷(万)": round(sd_loan / 10000, 2) if sd_loan > 0 else 0,
                    "商贷月供(元)": round(sd_m, 0) if sd_loan > 0 else 0,
                    "商贷总利息(万)": round(sd_interest / 10000, 2) if sd_loan > 0 else 0,
                    "合计月供(元)": round(gjj_m + sd_m, 0),
                    "合计总利息(万)": round((gjj_interest + sd_interest) / 10000, 2),
                }
                rows.append(row)

        df = pd.DataFrame(rows)

        # ── 动态列 ─────────────────────────────────────────────────────────────
        if loan_type == "纯商业贷款":
            display_cols = ["场景", "房屋总价(万)", "首付比例", "首付(万)",
                            "商贷(万)", "商贷月供(元)", "商贷总利息(万)", "合计月供(元)", "合计总利息(万)"]
        elif loan_type == "纯公积金贷款":
            display_cols = ["场景", "房屋总价(万)", "首付比例", "首付(万)",
                            "公积金贷款(万)", "公积金月供(元)", "公积金总利息(万)", "合计月供(元)", "合计总利息(万)"]
        else:
            display_cols = ["场景", "房屋总价(万)", "首付比例", "首付(万)",
                            "公积金贷款(万)", "公积金月供(元)", "公积金总利息(万)",
                            "商贷(万)", "商贷月供(元)", "商贷总利息(万)",
                            "合计月供(元)", "合计总利息(万)"]

        df_display = df[display_cols].copy()

        # ── 表格 ───────────────────────────────────────────────────────────────
        st.divider()
        st.header("📊 对比结果")

        def highlight_min(s):
            if s.name in ("合计月供(元)", "商贷月供(元)", "公积金月供(元)"):
                return ['background-color: #d4edda; font-weight: bold' if v == s.min() else '' for v in s]
            return ['' for _ in s]

        styled = df_display.style.apply(highlight_min, axis=0)
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # ── 可视化 ─────────────────────────────────────────────────────────────
        st.subheader("📈 图表对比")
        chart_df = df.copy()
        chart_df["场景"] = pd.Categorical(chart_df["场景"], categories=chart_df["场景"].tolist(), ordered=True)

        col_a, col_b = st.columns(2)

        with col_a:
            if loan_type == "组合贷款":
                fig1 = go.Figure()
                unique_dps = sorted(chart_df["首付比例"].unique().tolist(),
                                    key=lambda x: int(x.replace("%", "")))
                bases = px.colors.qualitative.Set2[:len(unique_dps)]
                for i, dp in enumerate(unique_dps):
                    base = bases[i % len(bases)]
                    sub = chart_df[chart_df["首付比例"] == dp]
                    fig1.add_bar(name=f"公积金 {dp}", x=sub["场景"], y=sub["公积金月供(元)"],
                                 text=sub["公积金月供(元)"].apply(lambda x: f"{x:,.0f}"),
                                 textposition="inside", marker_color=lighten(base, 0.45), legendgroup=dp)
                    fig1.add_bar(name=f"商贷 {dp}", x=sub["场景"], y=sub["商贷月供(元)"],
                                 text=sub["商贷月供(元)"].apply(lambda x: f"{x:,.0f}"),
                                 textposition="inside", marker_color=darken(base, 0.35), legendgroup=dp)
                fig1.update_layout(barmode="stack", title="月供构成对比（元）", height=400,
                                   legend=dict(orientation="h", y=1.15, traceorder="grouped"))
            else:
                y_col = "商贷月供(元)" if loan_type == "纯商业贷款" else "公积金月供(元)"
                name = "商贷月供" if loan_type == "纯商业贷款" else "公积金月供"
                fig1 = px.bar(chart_df, x="场景", y=y_col, title=f"{name}对比（元）",
                               text=chart_df[y_col].apply(lambda x: f"{x:,.0f}"),
                               color="首付比例", color_discrete_sequence=px.colors.qualitative.Set2)
                fig1.update_traces(textposition="outside")
                fig1.update_layout(height=400, legend=dict(orientation="h", y=1.15))
            st.plotly_chart(fig1, use_container_width=True)

        with col_b:
            if loan_type == "组合贷款":
                fig2 = go.Figure()
                unique_dps = sorted(chart_df["首付比例"].unique().tolist(),
                                    key=lambda x: int(x.replace("%", "")))
                bases = px.colors.qualitative.Set2[:len(unique_dps)]
                for i, dp in enumerate(unique_dps):
                    base = bases[i % len(bases)]
                    sub = chart_df[chart_df["首付比例"] == dp]
                    fig2.add_bar(name=f"公积金 {dp}", x=sub["场景"], y=sub["公积金总利息(万)"],
                                 text=sub["公积金总利息(万)"].apply(lambda x: f"{x:.1f}万"),
                                 textposition="inside", marker_color=lighten(base, 0.45), legendgroup=dp)
                    fig2.add_bar(name=f"商贷 {dp}", x=sub["场景"], y=sub["商贷总利息(万)"],
                                 text=sub["商贷总利息(万)"].apply(lambda x: f"{x:.1f}万"),
                                 textposition="inside", marker_color=darken(base, 0.35), legendgroup=dp)
                fig2.update_layout(barmode="stack", title="总利息构成对比（万元）", height=400,
                                   legend=dict(orientation="h", y=1.15, traceorder="grouped"))
            else:
                y_col = "商贷总利息(万)" if loan_type == "纯商业贷款" else "公积金总利息(万)"
                name = "商贷总利息" if loan_type == "纯商业贷款" else "公积金总利息"
                fig2 = px.bar(chart_df, x="场景", y=y_col, title=f"{name}对比（万元）",
                               text=chart_df[y_col].apply(lambda x: f"{x:.1f}万"),
                               color="首付比例", color_discrete_sequence=px.colors.qualitative.Set2)
                fig2.update_traces(textposition="outside")
                fig2.update_layout(height=400, legend=dict(orientation="h", y=1.15))
            st.plotly_chart(fig2, use_container_width=True)

        # 趋势线
        st.subheader("📉 月供趋势")
        trend_df = df.copy()
        trend_fig = px.line(
            trend_df, x="房屋总价(万)", y="合计月供(元)", color="首付比例",
            markers=True, title="不同首付比例下，月供随房价变化趋势",
            color_discrete_sequence=px.colors.qualitative.Set2,
        )
        trend_fig.update_traces(
            texttemplate=trend_df["合计月供(元)"].apply(lambda x: f"{x:,.0f}"),
            textposition="top center",
        )
        trend_fig.update_layout(height=420)
        st.plotly_chart(trend_fig, use_container_width=True)

        # ── Excel 导出 ─────────────────────────────────────────────────────────
        st.divider()
        st.subheader("📥 导出结果")
        excel_bytes = build_excel(df_display, loan_type, years)
        st.download_button(
            label="📥 导出 Excel 文件",
            data=excel_bytes,
            file_name=f"房贷计算结果_{loan_type}_{years}年.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # 脚注
        st.divider()
        st.caption("💡 采用**等额本息**还款方式计算，结果仅供参考，实际以银行审批为准。")


# ═══════════════════════════════════════════════════════════════════════════════
#  Tab 2: 首套二套购买顺序对比
# ═══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("🔄 哪一套先买更划算？")
    st.caption(
        "分别设置两套房产的总价、贷款方式和首付比例。"
        "切换购买顺序后，各套房的贷款方式/首付不变，仅商贷利率随首套/二套身份变化（3.0% vs 3.3%）。"
    )

    # ── 房产配置（左右并列） ─────────────────────────────────────────────
    st.markdown("### 🏠 房产配置")
    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("**房产 A**")
        price_a = st.number_input("总价（万元）", 1.0, value=200.0, step=1.0, format="%.0f", key="t2_pa")
        loan_a = st.selectbox("贷款方式", ["纯商业贷款", "纯公积金贷款", "组合贷款"], key="t2_loan_a")
        a_uses_gjj = loan_a in ("纯公积金贷款", "组合贷款")
        a_min_dp = 20 if a_uses_gjj else 15
        dp_a = st.slider("首付比例（%）", a_min_dp, 70, a_min_dp, 1, key="t2_dp_a")

    with col_b:
        st.markdown("**房产 B**")
        price_b = st.number_input("总价（万元）", 1.0, value=150.0, step=1.0, format="%.0f", key="t2_pb")
        loan_b = st.selectbox("贷款方式", ["纯商业贷款", "纯公积金贷款", "组合贷款"], key="t2_loan_b")
        b_uses_gjj = loan_b in ("纯公积金贷款", "组合贷款")
        b_min_dp = 20 if b_uses_gjj else 15
        dp_b = st.slider("首付比例（%）", b_min_dp, 70, b_min_dp, 1, key="t2_dp_b")

    # ── 共用设置 ─────────────────────────────────────────────────────────
    st.markdown("### ⚙️ 共用设置")
    c1, c2 = st.columns(2)
    with c1:
        t2_years = st.selectbox("贷款年限", YEAR_OPTIONS, index=5, key="t2_years")
    with c2:
        t2_cap = 80.0
        if a_uses_gjj or b_uses_gjj:
            t2_cap = st.number_input("公积金最高贷款额度（万元）", 0.0, value=80.0, step=1.0, format="%.0f", key="t2_cap",
                                     help="两套房使用相同公积金上限")

    # 单套房计算（贷款方式和首付比例跟随房产，利率跟随首套/二套身份）
    def _calc_one(price_wan, dp_pct, loan_mode, is_first):
        tp = price_wan * 10000
        down = tp * dp_pct / 100
        loan = tp - down
        cap = t2_cap * 10000
        sd_r = 3.0 if is_first else 3.3

        if loan_mode == "纯商业贷款":
            gjj, sd = 0.0, loan
        elif loan_mode == "纯公积金贷款":
            gjj, sd = min(loan, cap) if cap > 0 else loan, 0.0
        else:
            gjj = min(loan, cap) if cap > 0 else 0
            sd = loan - gjj

        gjj_m, gjj_i, _ = monthly_payment(gjj, gjj_rate, t2_years)
        sd_m, sd_i, _ = monthly_payment(sd, sd_r, t2_years)
        return dict(
            down=down/10000, gjj_loan=gjj/10000, gjj_m=gjj_m, gjj_i=gjj_i/10000,
            sd_loan=sd/10000, sd_m=sd_m, sd_i=sd_i/10000,
            total_m=gjj_m+sd_m, total_i=(gjj_i+sd_i)/10000,
            loan_mode=loan_mode, dp_pct=dp_pct,
        )

    # 方案一：先A（首套）→ B（二套）
    p1a = _calc_one(price_a, dp_a, loan_a, True)
    p1b = _calc_one(price_b, dp_b, loan_b, False)
    p1_total = p1a["total_i"] + p1b["total_i"]

    # 方案二：先B（首套）→ A（二套）
    p2b = _calc_one(price_b, dp_b, loan_b, True)
    p2a = _calc_one(price_a, dp_a, loan_a, False)
    p2_total = p2b["total_i"] + p2a["total_i"]

    # 展示
    st.divider()
    st.subheader("📊 两种方案对比")
    rc1, rc2 = st.columns(2)

    with rc1:
        better1 = " 🏆 更优" if p1_total < p2_total else ""
        st.markdown(f"**方案一：先A({price_a:.0f}万，{loan_a}，首付{dp_a}%，首套) → 后B({price_b:.0f}万，{loan_b}，首付{dp_b}%，二套){better1}**")
        st.markdown(f"""| | A（首套 3.0%） | B（二套 3.3%） | 合计 |
|---|---|---|---|
| 首付（万） | {p1a['down']:.2f} | {p1b['down']:.2f} | {p1a['down']+p1b['down']:.2f} |
| 商贷（万） | {p1a['sd_loan']:.2f} | {p1b['sd_loan']:.2f} | {p1a['sd_loan']+p1b['sd_loan']:.2f} |
| 合计月供（元） | {p1a['total_m']:,.0f} | {p1b['total_m']:,.0f} | {p1a['total_m']+p1b['total_m']:,.0f} |
| **总利息（万）** | **{p1a['total_i']:.2f}** | **{p1b['total_i']:.2f}** | **{p1_total:.2f}** |
""")

    with rc2:
        better2 = " 🏆 更优" if p2_total < p1_total else ""
        st.markdown(f"**方案二：先B({price_b:.0f}万，{loan_b}，首付{dp_b}%，首套) → 后A({price_a:.0f}万，{loan_a}，首付{dp_a}%，二套){better2}**")
        st.markdown(f"""| | B（首套 3.0%） | A（二套 3.3%） | 合计 |
|---|---|---|---|
| 首付（万） | {p2b['down']:.2f} | {p2a['down']:.2f} | {p2b['down']+p2a['down']:.2f} |
| 商贷（万） | {p2b['sd_loan']:.2f} | {p2a['sd_loan']:.2f} | {p2b['sd_loan']+p2a['sd_loan']:.2f} |
| 合计月供（元） | {p2b['total_m']:,.0f} | {p2a['total_m']:,.0f} | {p2b['total_m']+p2a['total_m']:,.0f} |
| **总利息（万）** | **{p2b['total_i']:.2f}** | **{p2a['total_i']:.2f}** | **{p2_total:.2f}** |
""")

    # 结论
    st.divider()
    st.subheader("🏆 结论")
    diff = abs(p1_total - p2_total)
    if diff < 0.01:
        st.info("两种方案总利息几乎相同，任选其一即可。")
    elif p1_total < p2_total:
        st.success(
            f"**先买A（{price_a:.0f}万）作为首套房更划算！** 可节省总利息约 **{diff:.2f} 万元**。\n\n"
            f"💡 总价更高的 A（{price_a:.0f}万）享受首套低利率（3.0%），"
            f"总价更低的 B（{price_b:.0f}万）承担二套高利率（3.3%），整体利息最低。"
        )
    else:
        st.success(
            f"**先买B（{price_b:.0f}万）作为首套房更划算！** 可节省总利息约 **{diff:.2f} 万元**。\n\n"
            f"💡 总价更高的 B（{price_b:.0f}万）享受首套低利率（3.0%），"
            f"总价更低的 A（{price_a:.0f}万）承担二套高利率（3.3%），整体利息最低。"
        )

    # 公积金明细
    t2_any_gjj = a_uses_gjj or b_uses_gjj
    if t2_any_gjj:
        with st.expander("🔍 查看公积金/商贷明细"):
            for label, first, second, first_mode, second_mode in [
                (f"方案一：先A→B", p1a, p1b, loan_a, loan_b),
                (f"方案二：先B→A", p2b, p2a, loan_b, loan_a),
            ]:
                st.caption(f"**{label}**")
                d1, d2 = st.columns(2)
                with d1:
                    st.write(
                        f"首套（{first_mode}）：公积金 {first['gjj_loan']:.2f}万"
                        f"（月供 {first['gjj_m']:,.0f} 利息 {first['gjj_i']:.2f}万）｜ "
                        f"商贷 {first['sd_loan']:.2f}万"
                        f"（月供 {first['sd_m']:,.0f} 利息 {first['sd_i']:.2f}万）"
                    )
                with d2:
                    st.write(
                        f"二套（{second_mode}）：公积金 {second['gjj_loan']:.2f}万"
                        f"（月供 {second['gjj_m']:,.0f} 利息 {second['gjj_i']:.2f}万）｜ "
                        f"商贷 {second['sd_loan']:.2f}万"
                        f"（月供 {second['sd_m']:,.0f} 利息 {second['sd_i']:.2f}万）"
                    )
